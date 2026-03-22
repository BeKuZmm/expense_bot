import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from database import get_expenses
from datetime import datetime
import os

def export_to_excel(user_id, period="month") -> str:
    rows = get_expenses(user_id, period)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Harajatlar"

    headers = ["#", "Miqdor (so'm)", "Kategoriya", "Izoh", "Sana"]
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total = 0
    for i, (amount, category, desc, date) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=amount)
        ws.cell(row=i, column=3, value=category)
        ws.cell(row=i, column=4, value=desc)
        ws.cell(row=i, column=5, value=date)
        total += amount
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = PatternFill(
                    start_color="F0F4F8", end_color="F0F4F8", fill_type="solid"
                )

    last_row = len(rows) + 2
    ws.cell(row=last_row, column=1, value="JAMI")
    ws.cell(row=last_row, column=2, value=total)
    for col in range(1, 3):
        ws.cell(row=last_row, column=col).font = Font(bold=True)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 18

    filename = f"harajatlar_{user_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename
